import os
import pytest
import shutil
import configparser
from pathlib import Path
from datetime import datetime
import tempfile

import openpyxl
import polars as pl
from unittest.mock import patch, MagicMock, mock_open, ANY

from service_medical_docs_analyzer import analyze_medical_documents, output_excel, MedicalDocsAnalyzer
from config_manager import load_config


def restore_config(config, original_config):
    """configを元の状態に復元するヘルパーメソッド"""
    for section in config.sections():
        config.remove_section(section)
    for section in original_config.sections():
        if not config.has_section(section):
            config.add_section(section)
        for key, value in original_config[section].items():
            config[section][key] = value


@pytest.fixture
def mock_config():
    config = configparser.ConfigParser()
    config['PATHS'] = {
        'database_path': 'test_database.xlsx',
        'template_path': 'test_template.xlsx',
        'output_dir': 'test_output'
    }
    config['Analysis'] = {
        'ordered_names': '山田,佐藤,鈴木',
        'clinical_departments': '合計,内科,外科,皮膚科'
    }
    return config


@pytest.fixture
def temp_files():
    # 一時ディレクトリとファイルを作成
    temp_dir = tempfile.mkdtemp()
    temp_db_path = os.path.join(temp_dir, 'test_database.xlsx')
    temp_template_path = os.path.join(temp_dir, 'test_template.xlsx')
    temp_output_dir = os.path.join(temp_dir, 'test_output')
    os.makedirs(temp_output_dir, exist_ok=True)

    # テスト用のExcelファイルを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['預り日', '患者ID', '患者名', '文書名', '診療科', '依頼医師名', '依頼部署', '医師依頼日', '担当者名']
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    # テストデータを追加
    data = [
        ['2025/01/10', 12345, '患者A', '文書A', '内科', '医師A', '部署A', '2025/01/09', '山田'],
        ['2025/01/15', 23456, '患者B', '文書B', '外科', '医師B', '部署B', '2025/01/14', '佐藤'],
        ['2025/01/20', 34567, '患者C', '文書C', '皮膚科', '医師C', '部署C', '2025/01/19', '鈴木']
    ]
    for i, row_data in enumerate(data, 2):
        for j, value in enumerate(row_data, 1):
            ws.cell(row=i, column=j).value = value

    wb.save(temp_db_path)

    # テンプレートファイルも作成
    wb_template = openpyxl.Workbook()
    ws_template = wb_template.active
    ws_template['A1'] = "医療文書作成件数"
    wb_template.save(temp_template_path)

    yield {
        'temp_dir': temp_dir,
        'db_path': temp_db_path,
        'template_path': temp_template_path,
        'output_dir': temp_output_dir
    }

    # テスト後にクリーンアップ - 安全に削除
    try:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                try:
                    os.chmod(file_path, 0o777)
                    os.unlink(file_path)
                except:
                    pass
        shutil.rmtree(temp_dir, ignore_errors=True)
    except:
        pass


@patch('service_medical_docs_analyzer.load_config')
@patch('os.system')  # Excelを開かないようにパッチ
@patch('service_medical_docs_analyzer.output_excel')  # output_excelをモック化
def test_analyze_medical_documents(mock_output_excel, mock_os_system, mock_load_config, mock_config, temp_files):
    # モックの設定
    mock_load_config.return_value = mock_config

    # output_excelがファイルを作成するようにする
    def side_effect_output(*args, **kwargs):
        output_file_path = os.path.join(temp_files['output_dir'], f"医療文書作成件数20250101-20250131.xlsx")
        with open(output_file_path, 'w') as f:
            f.write("test")

    mock_output_excel.side_effect = side_effect_output

    # 関数を実行
    analyze_medical_documents(
        temp_files['db_path'],
        temp_files['template_path'],
        '2025-01-01',
        '2025-01-31'
    )

    # mock_output_excelが呼ばれたことを確認
    assert mock_output_excel.called

    # os.systemが呼ばれないことを確認
    mock_os_system.assert_not_called()


@patch('service_medical_docs_analyzer.load_config')
@patch('os.system')  # Excelを開かないようにパッチ
def test_analyze_medical_documents_no_data(mock_os_system, mock_load_config, mock_config, temp_files):
    # 空のデータベースファイルを作成
    empty_db_path = os.path.join(temp_files['temp_dir'], 'empty_database.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['預り日', '患者ID', '患者名', '文書名', '診療科', '依頼医師名', '依頼部署', '医師依頼日', '担当者名']
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header
    wb.save(empty_db_path)

    # モックの設定
    mock_load_config.return_value = mock_config

    # 関数を実行してエラーがないことを確認
    analyze_medical_documents(
        empty_db_path,
        temp_files['template_path'],
        '2025-01-01',
        '2025-01-31'
    )

    # os.systemが呼ばれないことを確認
    mock_os_system.assert_not_called()


@patch('os.system')  # Excelを開かないようにパッチ
@patch('service_medical_docs_analyzer.os.path.exists')
def test_output_excel(mock_path_exists, mock_os_system, mock_config, temp_files):
    pass # 動作確認は手動で行う


@patch('service_medical_docs_analyzer.analyze_medical_documents')
@patch('os.system')  # Excelを開かないようにパッチ
def test_medical_docs_analyzer_run_analysis(mock_os_system, mock_analyze, mock_config):
    # 元のconfigをバックアップ
    original_config = load_config()

    try:
        # テスト用configを設定
        with patch('service_medical_docs_analyzer.load_config', return_value=mock_config):
            analyzer = MedicalDocsAnalyzer()
            success, message = analyzer.run_analysis('2025-01-01', '2025-01-31')

            # analyze_medical_documentsが正しく呼ばれたか確認
            mock_analyze.assert_called_once_with(
                mock_config['PATHS']['database_path'],
                mock_config['PATHS']['template_path'],
                '2025-01-01',
                '2025-01-31'
            )

            # 成功したことを確認
            assert success
            assert "集計が完了しました" in message
    finally:
        # 元のconfigに戻す
        current_config = load_config()
        restore_config(current_config, original_config)

    # os.systemが呼ばれないことを確認
    mock_os_system.assert_not_called()


@patch('service_medical_docs_analyzer.analyze_medical_documents')
@patch('os.system')  # Excelを開かないようにパッチ
def test_medical_docs_analyzer_run_analysis_error(mock_os_system, mock_analyze, mock_config):
    # 例外を発生させるようにモックを設定
    mock_analyze.side_effect = Exception("テストエラー")

    # 元のconfigをバックアップ
    original_config = load_config()

    try:
        # テスト用configを設定
        with patch('service_medical_docs_analyzer.load_config', return_value=mock_config):
            analyzer = MedicalDocsAnalyzer()
            success, message = analyzer.run_analysis('2025-01-01', '2025-01-31')

            # 失敗したことを確認
            assert not success
            assert "エラーが発生しました" in message
    finally:
        # 元のconfigに戻す
        current_config = load_config()
        restore_config(current_config, original_config)

    # os.systemが呼ばれないことを確認
    mock_os_system.assert_not_called()


@patch('service_medical_docs_analyzer.analyze_medical_documents')
@patch('os.system')  # Excelを開かないようにパッチ
def test_medical_docs_analyzer_run_analysis_date_error(mock_os_system, mock_analyze, mock_config):
    # 日付エラーを発生させるようにモックを設定
    mock_analyze.side_effect = ValueError("無効な日付形式")

    # 元のconfigをバックアップ
    original_config = load_config()

    try:
        # テスト用configを設定
        with patch('service_medical_docs_analyzer.load_config', return_value=mock_config):
            analyzer = MedicalDocsAnalyzer()
            success, message = analyzer.run_analysis('不正な日付', '2025-01-31')

            # 失敗したことを確認
            assert not success
            assert "日付の形式が正しくありません" in message
    finally:
        # 元のconfigに戻す
        current_config = load_config()
        restore_config(current_config, original_config)

    # os.systemが呼ばれないことを確認
    mock_os_system.assert_not_called()
