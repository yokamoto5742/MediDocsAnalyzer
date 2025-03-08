import pytest
import os
import shutil
import datetime
from pathlib import Path
from unittest.mock import patch, MagicMock, mock_open
import openpyxl
import polars as pl

from service_excel_handler import (
    backup_excel_file,
    get_last_row,
    apply_cell_formats,
    sort_worksheet_data,
    read_excel_to_dataframe,
    write_dataframe_to_excel
)


class TestBackupExcelFile:
    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.mkdir')
    @patch('shutil.copy2')
    def test_backup_excel_success(self, mock_copy2, mock_mkdir, mock_exists):
        # モックの設定
        mock_exists.return_value = False

        # テスト実行
        result = backup_excel_file('test.xlsx', 'backup_dir')

        # 検証
        mock_mkdir.assert_called_once_with(parents=True, exist_ok=True)
        mock_copy2.assert_called_once()
        assert 'backup_test.xlsx' in result

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.mkdir')
    @patch('shutil.copy2')
    def test_backup_excel_existing_dir(self, mock_copy2, mock_mkdir, mock_exists):
        # モックの設定
        mock_exists.return_value = True

        # テスト実行
        result = backup_excel_file('test.xlsx', 'backup_dir')

        # 検証
        mock_mkdir.assert_not_called()
        mock_copy2.assert_called_once()
        assert 'backup_test.xlsx' in result

    @patch('pathlib.Path.exists')
    @patch('shutil.copy2')
    def test_backup_excel_failure(self, mock_copy2, mock_exists):
        # モックの設定
        mock_exists.return_value = True
        mock_copy2.side_effect = Exception("テストエラー")

        # テスト実行
        result = backup_excel_file('test.xlsx', 'backup_dir')

        # 検証
        assert result is None


class TestGetLastRow:
    def test_get_last_row_with_data(self):
        # テスト用ワークシートの作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'テスト1'
        ws['B1'] = 'テスト2'
        ws['A2'] = 'データ1'
        ws['B2'] = 'データ2'
        ws['A3'] = 'データ3'

        # テスト実行
        result = get_last_row(ws)

        # 検証
        assert result == 3

    def test_get_last_row_empty(self):
        # 空のワークシートを作成
        wb = openpyxl.Workbook()
        ws = wb.active

        # テスト実行
        result = get_last_row(ws)

        # 検証
        assert result == 0


class TestApplyCellFormats:
    @patch('service_excel_handler.get_last_row')
    def test_apply_cell_formats(self, mock_get_last_row):
        # モックの設定
        mock_get_last_row.return_value = 3

        # テスト用ワークシートの作成
        wb = openpyxl.Workbook()
        ws = wb.active

        # テスト実行
        apply_cell_formats(ws, 2)

        # 検証 - セルに適切なフォーマットが適用されていることを確認
        assert ws.cell(row=2, column=1).alignment.horizontal == 'center'
        assert ws.cell(row=2, column=3).alignment.horizontal == 'left'
        assert ws.cell(row=2, column=3).alignment.shrink_to_fit == True


class TestSortWorksheetData:
    def test_sort_worksheet_data(self):
        # テスト用ワークシートの作成
        wb = openpyxl.Workbook()
        ws = wb.active

        # ヘッダー行
        ws['A1'] = '預り日'
        ws['B1'] = '患者ID'
        ws['E1'] = '診療科'

        # データ行
        ws['A2'] = datetime.datetime(2023, 5, 15)
        ws['B2'] = 102
        ws['E2'] = '内科'

        ws['A3'] = datetime.datetime(2023, 5, 10)
        ws['B3'] = 101
        ws['E3'] = '外科'

        ws['A4'] = datetime.datetime(2023, 5, 15)
        ws['B4'] = 103
        ws['E4'] = '内科'

        # テスト実行
        sort_worksheet_data(ws)

        # 検証 - ソート後のデータを確認
        assert ws['A2'].value == datetime.datetime(2023, 5, 10)
        assert ws['B2'].value == 101
        assert ws['E2'].value == '外科'

        assert ws['A3'].value == datetime.datetime(2023, 5, 15)
        assert ws['B3'].value == 102
        assert ws['E3'].value == '内科'

        assert ws['A4'].value == datetime.datetime(2023, 5, 15)
        assert ws['B4'].value == 103
        assert ws['E4'].value == '内科'

    def test_sort_worksheet_data_empty(self):
        # 空のワークシートを作成
        wb = openpyxl.Workbook()
        ws = wb.active

        # テスト実行 - エラーが発生しないことを確認
        sort_worksheet_data(ws)  # 例外が発生しなければ成功


class TestReadExcelToDataframe:
    @patch('openpyxl.load_workbook')
    def test_read_excel_success(self, mock_load_workbook):
        # モックの設定
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.active = mock_sheet
        mock_load_workbook.return_value = mock_wb

        # ヘッダー行のモック
        header_row = [MagicMock() for _ in range(9)]
        for i, cell in enumerate(header_row):
            cell.value = f"Header{i + 1}"
        mock_sheet.__getitem__.return_value.__getitem__.return_value = header_row

        # データ行のモック
        data_row1 = [MagicMock() for _ in range(9)]
        data_row2 = [MagicMock() for _ in range(9)]
        for i, cell in enumerate(data_row1):
            cell.value = f"Data1_{i + 1}"
        for i, cell in enumerate(data_row2):
            cell.value = f"Data2_{i + 1}"

        mock_sheet.iter_rows.return_value = [data_row1, data_row2]

        # テスト実行
        df, headers = read_excel_to_dataframe("test.xlsx")

        # 検証
        assert len(headers) == 9
        assert headers[0] == "Header1"
        assert df.height == 2

    @patch('openpyxl.load_workbook')
    def test_read_excel_with_process_func(self, mock_load_workbook):
        # モックの設定
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.active = mock_sheet
        mock_load_workbook.return_value = mock_wb

        # ヘッダー行のモック
        header_row = [MagicMock() for _ in range(9)]
        for i, cell in enumerate(header_row):
            cell.value = f"Header{i + 1}"
        mock_sheet.__getitem__.return_value.__getitem__.return_value = header_row

        # データ行のモック
        data_row = [MagicMock() for _ in range(9)]
        for i, cell in enumerate(data_row):
            cell.value = i + 1

        mock_sheet.iter_rows.return_value = [data_row]

        # テスト用の処理関数
        def process_func(cell):
            if cell.value is not None:
                return cell.value * 2
            return None

        # テスト実行
        df, headers = read_excel_to_dataframe("test.xlsx", process_func)

        # 検証
        assert len(headers) == 9
        assert df.height == 1
        for i in range(9):
            assert df[0, i] == (i + 1) * 2

    @patch('openpyxl.load_workbook')
    def test_read_excel_exception(self, mock_load_workbook):
        # モックの設定
        mock_load_workbook.side_effect = Exception("テストエラー")

        # テスト実行
        df, headers = read_excel_to_dataframe("test.xlsx")

        # 検証
        assert df.height == 0
        assert len(headers) == 0


class TestWriteDataframeToExcel:
    @patch('openpyxl.Workbook')
    @patch('openpyxl.load_workbook')
    @patch('os.path.exists')
    def test_write_new_excel(self, mock_exists, mock_load_workbook, mock_workbook):
        # モックの設定
        mock_exists.return_value = False
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.active = mock_sheet
        mock_workbook.return_value = mock_wb

        # テストデータ作成
        df = pl.DataFrame({
            'A': [1, 2],
            'B': ['X', 'Y']
        })
        headers = ['A', 'B']

        # テスト実行
        result = write_dataframe_to_excel(df, "test.xlsx", headers, create_new=True)

        # 検証
        assert result == True
        mock_sheet.cell.assert_called()
        mock_wb.save.assert_called_once_with("test.xlsx")

    @patch('openpyxl.load_workbook')
    @patch('os.path.exists')
    def test_write_existing_excel(self, mock_exists, mock_load_workbook):
        # モックの設定
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet.max_row = 5
        mock_wb.active = mock_sheet
        mock_load_workbook.return_value = mock_wb

        # テストデータ作成
        df = pl.DataFrame({
            'A': [1, 2],
            'B': ['X', 'Y']
        })
        headers = ['A', 'B']

        # テスト実行
        result = write_dataframe_to_excel(df, "test.xlsx", headers, create_new=False)

        # 検証
        assert result == True
        # セルのクリア処理が行われたことを確認
        assert mock_sheet.cell.call_count > 0
        mock_wb.save.assert_called_once_with("test.xlsx")

    @patch('openpyxl.Workbook')
    @patch('os.path.exists')
    def test_write_with_format_func(self, mock_exists, mock_workbook):
        # モックの設定
        mock_exists.return_value = False
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.active = mock_sheet
        mock_workbook.return_value = mock_wb

        # テストデータ作成
        df = pl.DataFrame({
            'A': [1, 2],
            'B': ['X', 'Y']
        })
        headers = ['A', 'B']

        # フォーマット関数
        def format_func(col_idx, value):
            if col_idx == 1:  # A列
                return value * 10
            return value

        # テスト実行
        result = write_dataframe_to_excel(
            df, "test.xlsx", headers, create_new=True,
            format_cells=False, format_func=format_func
        )

        # 検証
        assert result == True
        mock_sheet.cell.assert_called()
        mock_wb.save.assert_called_once_with("test.xlsx")

    @patch('openpyxl.Workbook')
    @patch('os.path.exists')
    def test_write_excel_exception(self, mock_exists, mock_workbook):
        # モックの設定
        mock_exists.return_value = False
        mock_workbook.side_effect = Exception("テストエラー")

        # テストデータ作成
        df = pl.DataFrame({
            'A': [1, 2],
            'B': ['X', 'Y']
        })
        headers = ['A', 'B']

        # テスト実行
        result = write_dataframe_to_excel(df, "test.xlsx", headers)

        # 検証
        assert result == False
