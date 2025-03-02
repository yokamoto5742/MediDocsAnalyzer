import os
import datetime
import time
from pathlib import Path

import polars as pl
import pyautogui
import win32com.client
import win32gui
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PyQt6.QtWidgets import QMessageBox

from config_manager import ConfigManager


def get_last_row(worksheet):
    last_row = 0
    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            break
        last_row += 1
    return last_row


def apply_cell_formats(worksheet, start_row):
    last_row = get_last_row(worksheet)

    # A列からF列までの範囲を設定
    for row in range(start_row, last_row + 1):
        for col in range(1, 7):
            cell = worksheet.cell(row=row, column=col)

            cell.alignment = Alignment(vertical='center')

            if col in [1, 2, 5, 6]:  # A,B列とE,F列
                cell.alignment = Alignment(horizontal='center')

            elif col in [3, 4]:  # C列とD列
                cell.alignment = Alignment(horizontal='left', shrink_to_fit=True)


def sort_excel_data(worksheet):
    try:
        last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row

        # ソートの範囲を設定
        sort_range = worksheet.Range(f"A2:I{last_row}")

        sort_range.Sort(
            Key1=worksheet.Range("A2"),  # A列（預り日）
            Order1=1,  # 1=昇順
            Key2=worksheet.Range("E2"),  # E列（診療科）
            Order2=1,  # 1=昇順
            Key3=worksheet.Range("B2"),  # B列（患者ID）
            Order3=1,  # 1=昇順
            Header=1,  # 1=ヘッダーあり
            OrderCustom=1,
            MatchCase=False,
            Orientation=1)

        return last_row

    except Exception as e:
        print(f"ソート中にエラーが発生しました: {str(e)}")
        raise


def bring_excel_to_front():
    # Excelの表示を最前面にする（最大3回まで試行）
    for _ in range(2):
        hwnd = win32gui.FindWindow("XLMAIN", None)
        if hwnd:
            win32gui.SetForegroundWindow(hwnd)
            return True
        time.sleep(0.1)
    return False


def write_data_to_excel(excel_path, df):
    if not os.path.exists(excel_path) or not excel_path.endswith('.xlsm'):
        print(f"Excelファイルが見つかりません: {excel_path}")
        return False

    try:
        wb = load_workbook(filename=excel_path, read_only=False, keep_vba=True)
    except PermissionError:
        QMessageBox.critical(None,
                            "エラー",
                            "Excelファイルが別のプロセスで開かれています。\nファイルを閉じてから再度実行してください。"
                            )
        return False
    
    ws = wb.active

    # 実際のデータが存在する最終行を取得
    last_row = get_last_row(ws)

    # 既存データを保持するセットを作成（A列からF列までの値をキーとして使用）
    existing_data = set()
    for row in range(2, last_row + 1):  # ヘッダー行をスキップ
        # 日付をYYYYMMDD形式の文字列として取得
        date_value = ws.cell(row=row, column=1).value
        if isinstance(date_value, datetime.datetime):
            date_str = date_value.strftime('%Y%m%d')
        else:
            date_str = str(date_value or '')

        # A列からF列までの値を取得（日付は数値形式で保持）
        row_data = (
            date_str,  # 日付を8桁の数値文字列として保持
            str(ws.cell(row=row, column=2).value or ''),
            str(ws.cell(row=row, column=3).value or ''),
            str(ws.cell(row=row, column=4).value or ''),
            str(ws.cell(row=row, column=5).value or ''),
            str(ws.cell(row=row, column=6).value or '')
        )
        existing_data.add(row_data)

    # CSVデータを文字列に変換
    temp_df = df.select([
        pl.col('*').cast(pl.String)
    ])
    data_to_write = temp_df.to_numpy().tolist()

    # 重複していないデータのみを抽出
    unique_data = []
    for row in data_to_write:
        # CSVの日付を8桁の数値文字列に変換
        csv_date = row[0]
        if isinstance(csv_date, str):
            try:
                # YYYY-MM-DD形式をYYYYMMDD形式に変換
                date_obj = datetime.datetime.strptime(csv_date, '%Y-%m-%d')
                date_str = date_obj.strftime('%Y%m%d')
            except ValueError:
                date_str = csv_date
        else:
            date_str = str(csv_date)

        # 比較用のタプルを作成
        row_data = (
            date_str,
            str(row[1] or ''),
            str(row[2] or ''),
            str(row[3] or ''),
            str(row[4] or ''),
            str(row[5] or '')
        )

        if row_data not in existing_data:
            unique_data.append(row)

    # 重複しないデータのみを書き込む
    for i, row in enumerate(unique_data):
        for j, value in enumerate(row):
            cell = ws.cell(row=last_row + 1 + i, column=j + 1)

            if j == 0:  # 日付列
                try:
                    date_value = datetime.datetime.strptime(value, '%Y-%m-%d')
                    cell.value = date_value
                    cell.number_format = 'yyyy/mm/dd'
                except valueError:
                    cell.value = value
            elif j == 1:  # 患者ID列
                try:
                    cell.value = int(value.replace(',', ''))
                    cell.number_format = '0'
                except ValueError:
                    cell.value = value
            else:
                cell.value = value if value is not None else ""

    apply_cell_formats(ws, last_row + 1)

    try:
        wb.save(excel_path)
        wb.close()
        return True
    except PermissionError:
        QMessageBox.critical(None,
                            "エラー",
                            "Excelファイルが別のプロセスで開かれているため、保存できません。\nファイルを閉じてから再度実行してください。"
                            )
        if 'wb' in locals():
            wb.close()
        return False


def open_and_sort_excel(excel_path):
    excel_path_str = str(Path(excel_path).resolve())
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    bring_excel_to_front()  # 最前面に表示
    workbook = excel.Workbooks.Open(excel_path_str)
    excel.WindowState = -4137  # xlMaximized
    workbook.Windows(1).Activate()

    try:
        worksheet = workbook.ActiveSheet
        sort_excel_data(worksheet)

        last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row  # データが存在する最後の行を特定する
        worksheet.Cells(last_row, 1).Select()

        config = ConfigManager()
        wait_time = config.get_share_button_wait_time()
        time.sleep(wait_time)
        share_x, share_y = config.get_share_button_position()
        pyautogui.click(share_x, share_y)

    except Exception as e:
        print(f"共有ボタンのクリックに失敗しました: {str(e)}")
    finally:
        # 操作が終わったらExcelは開いたままにする
        pyautogui.hotkey('win', 'down')  # ウィンドウを最小化
