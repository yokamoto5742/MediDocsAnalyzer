import os
from shutil import copyfile

import openpyxl
import polars as pl

from config_manager import load_config
from service_data_processor import filter_dataframe_by_date_range
from service_excel_handler import read_excel_to_dataframe


def analyze_medical_documents(file_path, excel_template_path, start_date_str=None, end_date_str=None):
    config = load_config()

    try:

        df, _ = read_excel_to_dataframe(file_path)

        # 必要なカラムの確認
        if df.height == 0 or '担当者名' not in df.columns or '診療科' not in df.columns or '預り日' not in df.columns:
            print("分析対象となるデータがありません。")
            return

        # 日付範囲の取得と処理
        date_result = filter_dataframe_by_date_range(df, start_date_str, end_date_str)
        df = date_result['df']
        start_date_display = date_result['start_date_display']
        end_date_display = date_result['end_date_display']
        file_date_range = date_result['file_date_range']

        # データの集計
        grouped = df.filter(
            pl.col('担当者名').is_not_null() & pl.col('診療科').is_not_null()
        ).group_by(['担当者名', '診療科']).agg(
            pl.len().alias('作成件数')
        )

        staff_totals = df.filter(
            pl.col('担当者名').is_not_null()
        ).group_by('担当者名').agg(
            pl.len().alias('作成件数')
        ).sort('担当者名')

        dept_totals = df.filter(
            pl.col('診療科').is_not_null()
        ).group_by('診療科').agg(
            pl.len().alias('作成件数')
        ).sort('作成件数', descending=True)

        # configからスタッフと診療科の情報を取得
        ordered_names_str = config['Analysis'].get('ordered_names', "")
        staff_members = [name.strip() for name in ordered_names_str.split(',')] if ordered_names_str else []

        config_departments_str = config['Analysis'].get('clinical_departments', "")
        departments = [dept.strip() for dept in config_departments_str.split(',')] if config_departments_str else []

        # 集計結果をExcelに出力
        output_excel(excel_template_path, staff_members, departments, grouped, staff_totals,
                     dept_totals, start_date_display, end_date_display, file_date_range)

    except FileNotFoundError:
        print(f"エラー: ファイル '{file_path}' が見つかりません。")
    except Exception as e:
        print(f"エラー: データの読み込み中に問題が発生しました - {str(e)}")


def output_excel(excel_template_path, staff_members, departments, grouped_data, staff_totals,
                 dept_totals, start_date, end_date, file_date_range):
    try:
        config = load_config()
        output_dir = config['PATHS']['output_dir']

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_file = os.path.join(output_dir, f"医療文書作成件数{file_date_range}.xlsx")

        if os.path.exists(excel_template_path):
            copyfile(excel_template_path, output_file)
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

        sheet['A1'] = f"医療文書作成件数 {start_date}-{end_date}"

        if len(departments) > 0:
            sheet['A2'] = "氏名"
            for col_idx, dept in enumerate(departments, 2):
                sheet.cell(row=2, column=col_idx).value = dept

        has_total_column = '合計' in departments
        total_column_idx = departments.index('合計') + 2 if has_total_column else 0

        for row_idx, staff in enumerate(staff_members, 3):
            sheet.cell(row=row_idx, column=1).value = staff

            staff_total = 0

            for col_idx, dept in enumerate(departments, 2):
                if dept == '合計':
                    continue

                filtered_data = grouped_data.filter(
                    (pl.col('担当者名') == staff) & (pl.col('診療科') == dept)
                )

                if filtered_data.height > 0:
                    count = filtered_data.select('作成件数').item()
                    sheet.cell(row=row_idx, column=col_idx).value = count
                    staff_total += count
                else:
                    sheet.cell(row=row_idx, column=col_idx).value = 0

            if has_total_column:
                sheet.cell(row=row_idx, column=total_column_idx).value = staff_total

        total_row = len(staff_members) + 3
        sheet.cell(row=total_row, column=1).value = "合計"

        for col_idx, dept in enumerate(departments, 2):
            if dept == '合計':
                continue

            dept_data = dept_totals.filter(pl.col('診療科') == dept)
            if dept_data.height > 0:
                sheet.cell(row=total_row, column=col_idx).value = dept_data.select('作成件数').item()
            else:
                sheet.cell(row=total_row, column=col_idx).value = 0

        total_docs = staff_totals.select('作成件数').sum().item()
        if has_total_column:
            sheet.cell(row=total_row, column=total_column_idx).value = total_docs

        workbook.save(output_file)
        os.system(f'start excel.exe "{output_file}"')

    except Exception as e:
        print(f"エラー: Excelファイルの出力中に問題が発生しました - {str(e)}")


class MedicalDocsAnalyzer:
    def __init__(self):
        self.config = load_config()
        self.paths_config = self.config['PATHS']

    def run_analysis(self, start_date_str, end_date_str):
        try:
            database_path = self.paths_config['database_path']
            template_path = self.paths_config['template_path']

            analyze_medical_documents(database_path, template_path, start_date_str, end_date_str)

            return True, "集計が完了しました。"

        except ValueError as ve:
            return False, f"日付の形式が正しくありません: {str(ve)}"
        except Exception as e:
            return False, f"分析中にエラーが発生しました: {str(e)}"
