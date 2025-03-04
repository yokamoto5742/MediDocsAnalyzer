import polars as pl
import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path
import os
import datetime


def get_last_row(worksheet):
    """
    ワークシートの最終行を取得する

    Args:
        worksheet: openpyxlのワークシートオブジェクト

    Returns:
        int: 最終行の行番号
    """
    last_row = 0
    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            break
        last_row += 1
    return last_row


def apply_cell_formats(worksheet, start_row):
    """
    セルの書式を適用する

    Args:
        worksheet: openpyxlのワークシートオブジェクト
        start_row: 書式適用を開始する行番号
    """
    last_row = get_last_row(worksheet)

    # A列からI列までの範囲を設定
    for row in range(start_row, last_row + 1):
        for col in range(1, 10):  # A-I列（1-9）
            cell = worksheet.cell(row=row, column=col)

            cell.alignment = Alignment(vertical='center')

            if col in [1, 2, 5, 6, 7, 8]:  # A, B, E, F, H列
                cell.alignment = Alignment(horizontal='center', vertical='center')

            elif col in [3, 4, 9]:  # C, D, I列
                cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)


# sort_worksheet_data関数の修正

def sort_worksheet_data(worksheet):
    """
    ワークシートのデータを並べ替える（openpyxlで実装）

    Args:
        worksheet: openpyxlのワークシートオブジェクト
    """
    # データ部分（ヘッダー以外）を取得
    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    # データがない場合は何もしない
    if not data_rows:
        return

    # 並べ替え：預り日(0列目)昇順 → 診療科(4列目)昇順 → 患者ID(1列目)昇順
    sorted_rows = sorted(data_rows, key=lambda x: (
        x[0] or datetime.datetime.min if isinstance(x[0], datetime.datetime) else str(x[0] or ""),  # 預り日
        x[4] or "",  # 診療科
        x[1] or 0  # 患者ID
    ))

    # 並べ替え後のデータを書き込み
    for i, row_data in enumerate(sorted_rows, start=2):
        for j, value in enumerate(row_data, start=1):
            worksheet.cell(row=i, column=j).value = value


def process_medical_documents(source_file, target_file):
    """
    バックアップファイルからデータを取り込み、重複行と空欄行を削除する

    Args:
        source_file (str): バックアップファイルのパス
        target_file (str): ターゲットデータベースファイルのパス
    """
    print(f"処理を開始: {source_file} から {target_file} へデータを取り込みます")

    # バックアップファイルを読み込む
    try:
        source_wb = openpyxl.load_workbook(source_file)
        source_sheet = source_wb.active

        # バックアップファイルのデータをリストに変換（A列からI列まで）
        if source_sheet.max_row > 0:  # シートが空でないことを確認
            headers = [cell.value for cell in source_sheet[1][0:9] if cell.value is not None]  # A-I列 (0-8)

            data = []
            for row in source_sheet.iter_rows(min_row=2, max_col=9):  # A-I列まで
                processed_row = []
                for idx, cell in enumerate(row):
                    # 日付セルの特別処理
                    if cell.column == 1 and cell.value:  # A列（預り日）
                        if isinstance(cell.value, (datetime.datetime, datetime.date)):
                            # 日付オブジェクトを文字列に変換
                            processed_row.append(cell.value.strftime('%Y/%m/%d'))
                        else:
                            processed_row.append(cell.value)
                    elif cell.column == 2 and cell.value:  # B列（患者ID）
                        # 患者IDを数値として処理
                        try:
                            if isinstance(cell.value, str):
                                # 文字列の場合は数値に変換
                                processed_row.append(int(cell.value))
                            else:
                                processed_row.append(cell.value)
                        except (ValueError, TypeError):
                            # 数値変換できない場合は元の値を使用
                            processed_row.append(cell.value)
                    elif cell.column == 8 and cell.value:  # H列（医師依頼日）
                        if isinstance(cell.value, (datetime.datetime, datetime.date)):
                            # 日付オブジェクトを文字列に変換
                            processed_row.append(cell.value.strftime('%Y/%m/%d'))
                        else:
                            processed_row.append(cell.value)
                    else:
                        processed_row.append(cell.value)
                data.append(processed_row)

            # Polarsデータフレームに変換
            df = pl.DataFrame(data, schema=headers, orient="row")

            print(f"バックアップファイルから {len(df)} 行のデータを読み込みました")

            # ターゲットファイルが存在する場合、そのデータを読み込む
            if os.path.exists(target_file):
                target_wb = openpyxl.load_workbook(target_file)
                target_sheet = target_wb.active

                if target_sheet.max_row > 0:  # シートが空でないことを確認
                    target_headers = [cell.value for cell in target_sheet[1][0:9] if
                                      cell.value is not None]  # A-I列 (0-8)

                    target_data = []
                    for row in target_sheet.iter_rows(min_row=2, max_col=9):  # A-I列まで
                        processed_row = []
                        for idx, cell in enumerate(row):
                            # 日付セルの特別処理
                            if cell.column == 1 and cell.value:  # A列（預り日）
                                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                                    # 日付オブジェクトを文字列に変換
                                    processed_row.append(cell.value.strftime('%Y/%m/%d'))
                                else:
                                    processed_row.append(cell.value)
                            elif cell.column == 2 and cell.value:  # B列（患者ID）
                                # 患者IDを数値として処理
                                try:
                                    if isinstance(cell.value, str):
                                        # 文字列の場合は数値に変換
                                        processed_row.append(int(cell.value))
                                    else:
                                        processed_row.append(cell.value)
                                except (ValueError, TypeError):
                                    # 数値変換できない場合は元の値を使用
                                    processed_row.append(cell.value)
                            elif cell.column == 8 and cell.value:  # H列（医師依頼日）
                                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                                    # 日付オブジェクトを文字列に変換
                                    processed_row.append(cell.value.strftime('%Y/%m/%d'))
                                else:
                                    processed_row.append(cell.value)
                            else:
                                processed_row.append(cell.value)
                        target_data.append(processed_row)

                    target_df = pl.DataFrame(target_data, schema=target_headers, orient="row")
                else:
                    # ターゲットシートが空の場合は空のデータフレームを作成
                    target_df = pl.DataFrame(schema=headers)

                print(f"ターゲットファイルから {len(target_df)} 行のデータを読み込みました")

                # スキーマの統一を確保
                # 両方のデータフレームのすべての列を文字列型に変換
                df = df.select([pl.col(col).cast(pl.Utf8) for col in df.columns])

                if len(target_df) > 0:
                    target_df = target_df.select([pl.col(col).cast(pl.Utf8) for col in target_df.columns])

                    # 列名が同じことを確保
                    if set(df.columns) == set(target_df.columns):
                        # 同じ列の順序で揃える
                        target_df = target_df.select(df.columns)
                        # データを結合
                        df = pl.concat([df, target_df])
                    else:
                        print(f"警告: ソースとターゲットのカラム構造が異なります。")
                        print(f"ソース: {df.columns}")
                        print(f"ターゲット: {target_df.columns}")
                        print("ソースファイルのデータのみを使用します。")

            # すべての列を文字列型に変換し、空のセルを空文字に
            for col in df.columns:
                df = df.with_columns([
                    pl.col(col).cast(pl.Utf8).fill_null("").alias(col)
                ])

            # 医師依頼日が空欄の行を削除
            if "医師依頼日" in df.columns:
                # 文字列型に変換してから比較
                df = df.with_columns([
                    pl.col("医師依頼日").cast(pl.Utf8).fill_null("").alias("医師依頼日")
                ])
                df = df.filter(pl.col("医師依頼日") != "")
            else:
                print("警告: '医師依頼日'の列が見つかりません。この条件でのフィルタリングをスキップします。")

            print(f"空の医師依頼日を持つ行を削除した後: {len(df)} 行")

            # 担当者名が空欄の行を削除
            if "担当者名" in df.columns:
                # 文字列型に変換してから比較
                df = df.with_columns([
                    pl.col("担当者名").cast(pl.Utf8).fill_null("").alias("担当者名")
                ])
                df = df.filter(pl.col("担当者名") != "")
                print(f"空の担当者名を持つ行を削除した後: {len(df)} 行")
            else:
                print("警告: '担当者名'の列が見つかりません。この条件でのフィルタリングをスキップします。")

            # 重複行を削除（預り日、患者ID、文書名、診療科、医師名の組み合わせが同じ行）
            # まず必要な列が存在するか確認
            required_columns = ["預り日", "患者ID", "文書名", "診療科", "医師名"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if not missing_columns:
                df = df.unique(subset=required_columns)
            else:
                print(f"警告: 次の列が見つからないため、重複削除をスキップします: {missing_columns}")
                # 存在する列のみで重複削除を試みる
                existing_columns = [col for col in required_columns if col in df.columns]
                if existing_columns:
                    print(f"代わりに次の列で重複削除を行います: {existing_columns}")
                    df = df.unique(subset=existing_columns)

            print(f"重複削除後: {len(df)} 行")

            if os.path.exists(target_file):
                try:
                    result_wb = openpyxl.load_workbook(target_file)
                    result_sheet = result_wb.active

                    # 既存のデータをクリア (セルの値のみを消去し、書式は保持)
                    data_rows = result_sheet.max_row
                    data_cols = 9  # A-I列まで
                    for row in range(2, data_rows + 1):  # ヘッダー以外をクリア
                        for col in range(1, data_cols + 1):
                            cell = result_sheet.cell(row=row, column=col)
                            cell.value = None  # 値のみクリア
                except Exception as e:
                    print(f"既存ファイルを開く際にエラーが発生しました: {e}")
                    # エラーの場合は新規作成
                    result_wb = openpyxl.Workbook()
                    result_sheet = result_wb.active
            else:
                result_wb = openpyxl.Workbook()
                result_sheet = result_wb.active

            # ヘッダーを書き込み（A-I列）
            for col_idx, header in enumerate(headers, 1):
                if col_idx <= 9:  # A-I列まで
                    result_sheet.cell(row=1, column=col_idx, value=header)

            # データを書き込み（A-I列）
            df_rows = df.rows()
            for row_idx, row_data in enumerate(df_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx <= 9:  # A-I列まで
                        cell = result_sheet.cell(row=row_idx, column=col_idx)

                        # 1列目（預り日）または8列目（医師依頼日）の場合、日付形式を適用
                        if (col_idx == 1 or col_idx == 8) and value:
                            # 日付文字列の処理（タイムスタンプ部分を取り除く）
                            if isinstance(value, str) and value:
                                # 日付部分のみを抽出（YYYY-MM-DD または YYYY/MM/DD 形式）
                                date_parts = value.split()[0] if ' ' in value else value

                                # 年月日の区切りを"/"に統一
                                if '-' in date_parts:
                                    parts = date_parts.split('-')
                                    if len(parts) == 3:
                                        year, month, day = parts
                                        formatted_date = f"{year}/{month}/{day}"
                                    else:
                                        formatted_date = date_parts
                                else:
                                    formatted_date = date_parts

                                cell.value = formatted_date
                                # 日付形式を設定
                                cell.number_format = 'yyyy/mm/dd'
                        # 2列目（患者ID）の場合、数値形式を適用
                        elif col_idx == 2 and value is not None:
                            # 数値として設定
                            try:
                                if isinstance(value, str) and value.strip():
                                    cell.value = int(value)
                                else:
                                    cell.value = value
                                # 数値形式を設定（カンマなし）
                                cell.number_format = '0'
                            except (ValueError, TypeError):
                                cell.value = value
                        else:
                            cell.value = value

            # データを並べ替え
            print("データを並べ替えます...")
            sort_worksheet_data(result_sheet)

            # セルの書式を適用
            print("セルの書式を適用します...")
            apply_cell_formats(result_sheet, 2)  # 2行目（データ行の開始）から適用

            # 不要なデフォルトシートを削除（新規作成時のみ）
            if 'Sheet' in result_wb.sheetnames and len(result_wb.sheetnames) > 1:
                del result_wb['Sheet']

            # ファイルを保存
            result_wb.save(target_file)

            print(f"処理完了: {len(df)} 行のデータを {target_file} に保存しました")
            return True
        else:
            print("エラー: ソースシートにデータがありません")
            return False

    except Exception as e:
        print(f"エラーが発生しました: {e}")
        return False


# メイン実行
if __name__ == "__main__":
    source_file = "backup_医療文書担当一覧.xlsm"
    target_file = "医療文書担当一覧データベース.xlsx"

    success = process_medical_documents(source_file, target_file)

    if success:
        print("データ処理が正常に完了しました")
    else:
        print("データ処理中にエラーが発生しました")
