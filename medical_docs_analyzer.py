import openpyxl
import polars as pl
import warnings
from datetime import datetime
import os
from shutil import copyfile

from config_manager import load_config, get_ordered_names


def analyze_medical_documents(file_path, excel_template_path):

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        data = []

        for row in list(sheet.rows)[1:]:
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_data[headers[i]] = cell.value

            if row_data.get('預り日') is not None:
                data.append(row_data)

        df = pl.DataFrame(data)

        if df.height == 0 or '担当者名' not in df.columns or '診療科' not in df.columns or '預り日' not in df.columns:
            print("分析対象となるデータがありません。")
            return

        df = df.with_columns(
            pl.when(pl.col('預り日').is_not_null())
            .then(pl.col('預り日'))
            .otherwise(None)
            .alias('預り日')
        )

        valid_dates = df.filter(pl.col('預り日').is_not_null()).select('預り日')
        if valid_dates.height == 0:
            print("有効な日付データがありません。")
            return

        # 日付データが文字列の場合は、datetime型に変換してから処理
        min_date = valid_dates.select(pl.col('預り日').min()).item()
        max_date = valid_dates.select(pl.col('預り日').max()).item()

        # 日付フォーマットの処理（文字列か日付型かを判断）
        if isinstance(min_date, str):
            # 文字列の場合はdatetimeに変換
            try:
                min_date_obj = datetime.strptime(min_date, '%Y/%m/%d')
                max_date_obj = datetime.strptime(max_date, '%Y/%m/%d')
                start_date = min_date_obj.strftime('%Y年%m月%d日')
                end_date = max_date_obj.strftime('%Y年%m月%d日')
                file_date_range = f"{min_date_obj.strftime('%Y%m%d')}-{max_date_obj.strftime('%Y%m%d')}"
            except ValueError:
                # 日付形式が違う場合はそのまま使用
                start_date = min_date
                end_date = max_date
                file_date_range = f"{start_date}-{end_date}".replace('/', '')
        else:
            # datetimeオブジェクトの場合はstrftimeを使用
            start_date = min_date.strftime('%Y年%m月%d日')
            end_date = max_date.strftime('%Y年%m月%d日')
            file_date_range = f"{min_date.strftime('%Y%m%d')}-{max_date.strftime('%Y%m%d')}"

        # 担当者と診療科で集計
        grouped = df.filter(
            pl.col('担当者名').is_not_null() & pl.col('診療科').is_not_null()
        ).group_by(['担当者名', '診療科']).agg(
            pl.len().alias('作成件数')
        )

        # 担当者ごとの合計件数を計算
        staff_totals = df.filter(
            pl.col('担当者名').is_not_null()
        ).group_by('担当者名').agg(
            pl.len().alias('作成件数')
        ).sort('担当者名')

        # 診療科ごとの合計件数を計算
        dept_totals = df.filter(
            pl.col('診療科').is_not_null()
        ).group_by('診療科').agg(
            pl.len().alias('作成件数')
        ).sort('作成件数', descending=True)

        # 担当者のリストを取得（None値を除外）
        staff_members = df.select('担当者名').filter(
            pl.col('担当者名').is_not_null()
        ).unique().sort('担当者名').to_series().to_list()

        ordered_names_str = config['Analysis'].get('config_ordered_names', "")
        config_ordered_names = [name.strip() for name in ordered_names_str.split(',')] if ordered_names_str else []

        # リストを指定された順序に並べ替え（データにない名前は無視）
        staff_members = [name for name in config_ordered_names if name in staff_members]

        # 診療科のリストを取得（None値を除外）
        departments = df.select('診療科').filter(
            pl.col('診療科').is_not_null()
        ).unique().sort('診療科').to_series().to_list()

        output_excel(excel_template_path, staff_members, departments, grouped, staff_totals,
                     dept_totals, start_date, end_date, file_date_range)

    except FileNotFoundError:
        print(f"エラー: ファイル '{file_path}' が見つかりません。")
    except Exception as e:
        print(f"エラー: データの読み込み中に問題が発生しました - {str(e)}")


def output_excel(excel_template_path, staff_members, departments, grouped_data, staff_totals,
                 dept_totals, start_date, end_date, file_date_range):
    try:
        output_file = f"医療文書作成件数{file_date_range}.xlsx"

        # テンプレートが存在する場合はコピー、存在しない場合は新規作成
        if os.path.exists(excel_template_path):
            copyfile(excel_template_path, output_file)
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        else:
            print(f"警告: テンプレートファイル '{excel_template_path}' が見つかりません。新規ファイルを作成します。")
            workbook = openpyxl.Workbook()
            sheet = workbook.active

        # タイトル行の設定
        sheet['A1'] = f"医療文書作成件数 {start_date}-{end_date}"

        # ヘッダー行の設定（すでにテンプレートにある場合は上書き）
        if len(departments) > 0:
            sheet['A2'] = "氏名"
            for col_idx, dept in enumerate(departments, 2):
                sheet.cell(row=2, column=col_idx).value = dept
            sheet.cell(row=2, column=len(departments) + 2).value = "合計"

        # データの書き込み
        for row_idx, staff in enumerate(staff_members, 3):
            # 担当者名
            sheet.cell(row=row_idx, column=1).value = staff

            # 各診療科の件数
            staff_total = 0
            for col_idx, dept in enumerate(departments, 2):
                # この担当者×診療科の件数を取得
                filtered_data = grouped_data.filter(
                    (pl.col('担当者名') == staff) & (pl.col('診療科') == dept)
                )

                if filtered_data.height > 0:
                    count = filtered_data.select('作成件数').item()
                    sheet.cell(row=row_idx, column=col_idx).value = count
                    staff_total += count
                else:
                    sheet.cell(row=row_idx, column=col_idx).value = 0

            # 合計
            sheet.cell(row=row_idx, column=len(departments) + 2).value = staff_total

        # 合計行
        total_row = len(staff_members) + 3
        sheet.cell(row=total_row, column=1).value = "合計"

        # 各診療科の合計
        for col_idx, dept in enumerate(departments, 2):
            dept_data = dept_totals.filter(pl.col('診療科') == dept)
            if dept_data.height > 0:
                sheet.cell(row=total_row, column=col_idx).value = dept_data.select('作成件数').item()
            else:
                sheet.cell(row=total_row, column=col_idx).value = 0

        # 全体の合計
        total_docs = staff_totals.select('作成件数').sum().item()
        sheet.cell(row=total_row, column=len(departments) + 2).value = total_docs

        # ファイルを保存
        workbook.save(output_file)
        print(f"集計結果を '{output_file}' に保存しました。")

    except Exception as e:
        print(f"エラー: Excelファイルの出力中に問題が発生しました - {str(e)}")


# 使用例
if __name__ == "__main__":
    config = load_config()
    ordered_names = get_ordered_names(config)
    database_path = config['PATHS']['database_path']
    template_path = config['PATHS']['template_path']
    analyze_medical_documents(database_path, template_path)
