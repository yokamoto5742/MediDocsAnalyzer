import os
import shutil
from pathlib import Path
import datetime

import openpyxl
from openpyxl.styles import Alignment
import polars as pl


def backup_excel_file(file_path, backup_dir):
    try:
        backup_dir_path = Path(backup_dir)
        if not backup_dir_path.exists():
            backup_dir_path.mkdir(parents=True, exist_ok=True)

        file_name = Path(file_path).name
        backup_file_name = f"backup_{file_name}"
        backup_path = backup_dir_path / backup_file_name

        shutil.copy2(file_path, backup_path)
        return str(backup_path)
    except Exception as e:
        print(f"バックアップ作成中にエラーが発生しました: {str(e)}")
        return None


def get_last_row(worksheet):
    last_row = 0
    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            break
        last_row += 1
    return last_row


def apply_cell_formats(worksheet, start_row):
    last_row = get_last_row(worksheet)

    # A列からI列までの範囲を設定
    for row in range(start_row, last_row + 1):
        for col in range(1, 10):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center')

            if col in [1, 2, 5, 6, 7, 8]:  # A, B, E, F, H列
                cell.alignment = Alignment(horizontal='center', vertical='center')

            elif col in [3, 4, 9]:  # C, D, I列
                cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)


def sort_worksheet_data(worksheet):
    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    if not data_rows:
        return

    sorted_rows = sorted(data_rows, key=lambda x: (
        x[0] or datetime.datetime.min if isinstance(x[0], datetime.datetime) else str(x[0] or ""),  # 預り日
        x[4] or "",  # 診療科
        x[1] or 0  # 患者ID
    ))

    # 並べ替え後のデータを書き込み
    for i, row_data in enumerate(sorted_rows, start=2):
        for j, value in enumerate(row_data, start=1):
            worksheet.cell(row=i, column=j).value = value


def read_excel_to_dataframe(file_path, process_cell_func=None):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1][0:9] if cell.value is not None]  # A-I列

        data = []
        for row in sheet.iter_rows(min_row=2, max_col=9):
            if process_cell_func:
                processed_row = [process_cell_func(cell) for cell in row]
            else:
                processed_row = [cell.value for cell in row]
            data.append(processed_row)

        return pl.DataFrame(data, schema=headers, orient="row"), headers
    except Exception as e:
        print(f"Excelファイルの読み込み中にエラーが発生しました: {str(e)}")
        return pl.DataFrame(), []


def write_dataframe_to_excel(df, file_path, headers, create_new=False, format_cells=True, format_func=None):
    try:
        if create_new or not os.path.exists(file_path):
            result_wb = openpyxl.Workbook()
            result_sheet = result_wb.active
        else:
            result_wb = openpyxl.load_workbook(file_path)
            result_sheet = result_wb.active

            # 既存のデータをクリア (セルの値のみを消去し、書式は保持)
            data_rows = result_sheet.max_row
            data_cols = 9  # A-I列まで
            for row in range(2, data_rows + 1):  # ヘッダー以外をクリア
                for col in range(1, data_cols + 1):
                    cell = result_sheet.cell(row=row, column=col)
                    cell.value = None  # 値のみクリア

        # ヘッダーを書き込み（A-I列）
        for col_idx, header in enumerate(headers, 1):
            if col_idx <= 9:  # A-I列まで
                result_sheet.cell(row=1, column=col_idx, value=header)

        # データを書き込み（A-I列）
        df_rows = df.rows()
        for row_idx, row_data in enumerate(df_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                if col_idx <= 9:  # A-I列まで
                    cell = result_sheet.cell(row=row_idx, column=col_idx)
                    
                    if format_func:
                        cell.value = format_func(col_idx, value)
                    else:
                        cell.value = value

        if format_cells:
            sort_worksheet_data(result_sheet)
            apply_cell_formats(result_sheet, 2)  # 2行目（データ行の開始）から適用

        result_wb.save(file_path)
        return True
    except Exception as e:
        print(f"Excelファイルの書き込み中にエラーが発生しました: {str(e)}")
        return False
