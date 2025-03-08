import os
from datetime import datetime
from shutil import copyfile

import openpyxl
import polars as pl
import warnings

from config_manager import get_ordered_names, load_config


def analyze_medical_documents(file_path, excel_template_path, start_date_str=None, end_date_str=None):
    config = load_config()

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        data = []

        for row in list(sheet.rows)[1:]:
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_data[headers[i]] = cell.value

            if row_data.get('預り日') is not None:
                data.append(row_data)

        df = pl.DataFrame(data)

        if df.height == 0 or '担当者名' not in df.columns or '診療科' not in df.columns or '預り日' not in df.columns:
            print("分析対象となるデータがありません。")
            return

        df = df.with_columns(
            pl.when(pl.col('預り日').is_not_null())
            .then(pl.col('預り日'))
            .otherwise(None)
            .alias('預り日')
        )

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

            df = df.with_columns([
                pl.when(pl.col('預り日').cast(pl.Utf8).str.contains('-'))
                .then(pl.col('預り日').cast(pl.Utf8).str.replace('-', '/'))
                .otherwise(pl.col('預り日'))
                .alias('預り日')
            ])

            df = df.filter(
                (pl.col('預り日') >= start_date.strftime('%Y/%m/%d')) &
                (pl.col('預り日') <= end_date.strftime('%Y/%m/%d'))
            )

            valid_dates = df.filter(pl.col('預り日').is_not_null()).select('預り日')
            min_date = valid_dates.select(pl.col('預り日').min()).item()
            max_date = valid_dates.select(pl.col('預り日').max()).item()

            if isinstance(min_date, str):
                # 文字列の場合はdatetimeに変換
                try:
                    min_date_obj = datetime.strptime(min_date, '%Y/%m/%d')
                    max_date_obj = datetime.strptime(max_date, '%Y/%m/%d')
                    actual_start_date = min_date_obj.strftime('%Y%m%d')
                    actual_end_date = max_date_obj.strftime('%Y%m%d')
                    start_date_display = min_date_obj.strftime('%Y年%m月%d日')
                    end_date_display = max_date_obj.strftime('%Y年%m月%d日')
                except ValueError:
                    actual_start_date = min_date.replace('/', '')
                    actual_end_date = max_date.replace('/', '')
                    start_date_display = min_date
                    end_date_display = max_date
            else:
                actual_start_date = min_date.strftime('%Y%m%d')
                actual_end_date = max_date.strftime('%Y%m%d')
                start_date_display = min_date.strftime('%Y年%m月%d日')
                end_date_display = max_date.strftime('%Y年%m月%d日')

            file_date_range = f"{actual_start_date}-{actual_end_date}"
        else:
            valid_dates = df.filter(pl.col('預り日').is_not_null()).select('預り日')
            min_date = valid_dates.select(pl.col('預り日').min()).item()
            max_date = valid_dates.select(pl.col('預り日').max()).item()

            if isinstance(min_date, str):
                try:
                    min_date_obj = datetime.strptime(min_date, '%Y/%m/%d')
                    max_date_obj = datetime.strptime(max_date, '%Y/%m/%d')
                    start_date_display = min_date_obj.strftime('%Y年%m月%d日')
                    end_date_display = max_date_obj.strftime('%Y年%m月%d日')
                    file_date_range = f"{min_date_obj.strftime('%Y%m%d')}-{max_date_obj.strftime('%Y%m%d')}"
                except ValueError:
                    start_date_display = min_date
                    end_date_display = max_date
                    file_date_range = f"{start_date_display}-{end_date_display}".replace('/', '')
            else:
                start_date_display = min_date.strftime('%Y年%m月%d日')
                end_date_display = max_date.strftime('%Y年%m月%d日')
                file_date_range = f"{min_date.strftime('%Y%m%d')}-{max_date.strftime('%Y%m%d')}"

        grouped = df.filter(
            pl.col('担当者名').is_not_null() & pl.col('診療科').is_not_null()
        ).group_by(['担当者名', '診療科']).agg(
            pl.len().alias('作成件数')
        )

        staff_totals = df.filter(
            pl.col('担当者名').is_not_null()
        ).group_by('担当者名').agg(
            pl.len().alias('作成件数')
        ).sort('担当者名')

        dept_totals = df.filter(
            pl.col('診療科').is_not_null()
        ).group_by('診療科').agg(
            pl.len().alias('作成件数')
        ).sort('作成件数', descending=True)

        ordered_names_str = config['Analysis'].get('ordered_names', "")
        staff_members = [name.strip() for name in ordered_names_str.split(',')] if ordered_names_str else []

        config_departments_str = config['Analysis'].get('clinical_departments', "")
        departments = [dept.strip() for dept in config_departments_str.split(',')] if config_departments_str else []

        output_excel(excel_template_path, staff_members, departments, grouped, staff_totals,
                     dept_totals, start_date_display, end_date_display, file_date_range)

    except FileNotFoundError:
        print(f"エラー: ファイル '{file_path}' が見つかりません。")
    except Exception as e:
        print(f"エラー: データの読み込み中に問題が発生しました - {str(e)}")


def output_excel(excel_template_path, staff_members, departments, grouped_data, staff_totals,
                 dept_totals, start_date, end_date, file_date_range):
    try:
        config = load_config()
        output_dir = config['PATHS']['output_dir']

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_file = os.path.join(output_dir, f"医療文書作成件数{file_date_range}.xlsx")

        if os.path.exists(excel_template_path):
            copyfile(excel_template_path, output_file)
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

        sheet['A1'] = f"医療文書作成件数 {start_date}-{end_date}"

        if len(departments) > 0:
            sheet['A2'] = "氏名"
            for col_idx, dept in enumerate(departments, 2):
                sheet.cell(row=2, column=col_idx).value = dept

        for row_idx, staff in enumerate(staff_members, 3):
            sheet.cell(row=row_idx, column=1).value = staff

            staff_total = 0
            has_total_column = '合計' in departments
            total_column_idx = departments.index('合計') + 2 if has_total_column else 0

            for col_idx, dept in enumerate(departments, 2):
                if dept == '合計':
                    # 合計列は後で設定するのでスキップ
                    continue

                filtered_data = grouped_data.filter(
                    (pl.col('担当者名') == staff) & (pl.col('診療科') == dept)
                )

                if filtered_data.height > 0:
                    count = filtered_data.select('作成件数').item()
                    sheet.cell(row=row_idx, column=col_idx).value = count
                    staff_total += count
                else:
                    sheet.cell(row=row_idx, column=col_idx).value = 0


            if has_total_column:
                sheet.cell(row=row_idx, column=total_column_idx).value = staff_total

        total_row = len(staff_members) + 3
        sheet.cell(row=total_row, column=1).value = "合計"

        for col_idx, dept in enumerate(departments, 2):
            if dept == '合計':
                continue

            dept_data = dept_totals.filter(pl.col('診療科') == dept)
            if dept_data.height > 0:
                sheet.cell(row=total_row, column=col_idx).value = dept_data.select('作成件数').item()
            else:
                sheet.cell(row=total_row, column=col_idx).value = 0

        total_docs = staff_totals.select('作成件数').sum().item()
        if has_total_column:
            sheet.cell(row=total_row, column=total_column_idx).value = total_docs

        workbook.save(output_file)
        os.system(f'start excel.exe "{output_file}"')

    except Exception as e:
        print(f"エラー: Excelファイルの出力中に問題が発生しました - {str(e)}")
